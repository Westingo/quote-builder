"""
build_codes.py — regenerate codes.yaml from Metro's Copy_Paste master workbook.

Usage:  python build_codes.py [path/to/Copy_Paste.xlsx]

Reads the 14 dictionary sheets, skips the reference sheets (MyQ Price Tool,
Labor Rates, Do not change), and emits codes.yaml keyed by sheet. Cost/labor are
carried for reference only — they are often freeform text, never assume numbers.
"""
import re
import sys
from openpyxl import load_workbook
import yaml

SRC = sys.argv[1] if len(sys.argv) > 1 else "Copy_Paste.xlsx"
OUT = "codes.yaml"

DICT_SHEETS = ['Brivo', 'Access', 'Gate Accessories', 'Operators', 'Ironwork',
               'CCTV', 'Remotes', 'Other', 'Readers', 'Door Hardware', 'Storage',
               'N', 'W', 'E']
SECTION = {'N': 'note', 'W': 'warranty', 'E': 'exclusion'}
KNOWN = {'CODE', 'DESCRIPTION', 'MODEL', 'COST', 'LABOR', 'MISC.'}


def coerce_code(v):
    """Normalize a code cell into a clean single-token key. Some workbook cells
    carry an annotation, e.g. 'CAPXLV\\n(Use with CAPXLV2)' — strip parenthetical
    notes and keep only the first line so the lookup key is just the code
    (otherwise the code imports un-findable, like the old CAPXLV entry)."""
    if v is None:
        return None
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    s = re.sub(r"\([^)]*\)", " ", str(v))   # drop (…) annotations
    s = s.split("\n")[0]                     # keep the first line only
    s = " ".join(s.split()).strip()          # collapse whitespace / newlines
    return s or None


def coerce_val(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def main():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    out = {}
    for name in DICT_SHEETS:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        # locate the header row (first row containing CODE and DESCRIPTION)
        hidx = None
        header = []
        for i, r in enumerate(rows[:8]):
            up = [str(c).strip().upper() if c is not None else '' for c in r]
            if 'CODE' in up and 'DESCRIPTION' in up:
                hidx, header = i, up
                break
        if hidx is None:
            print(f"  WARNING: no header found on sheet {name!r}, skipping")
            continue

        col = {}
        for j, h in enumerate(header):
            if h in KNOWN and h not in col:
                col[h] = j
        notecols = [j for j, h in enumerate(header) if h not in KNOWN]

        items, category = [], None
        for r in rows[hidx + 1:]:
            get = lambda key: r[col[key]] if key in col and col[key] < len(r) else None
            code = coerce_code(get('CODE'))
            desc = coerce_val(get('DESCRIPTION'))
            model = coerce_val(get('MODEL'))
            cost = coerce_val(get('COST'))
            labor = coerce_val(get('LABOR'))
            notes = [coerce_val(r[j]) for j in notecols if j < len(r)]
            notes = [n for n in notes if n not in (None, '')]
            if 'MISC.' in col:
                m = coerce_val(get('MISC.'))
                if m:
                    notes.insert(0, m)

            if all(x in (None, '') for x in [code, desc, model, cost, labor]) and not notes:
                continue
            # ALL-CAPS separator row: only the code cell populated
            if code and desc in (None, '') and model in (None, '') \
                    and cost in (None, '') and labor in (None, '') and not notes:
                category = code
                continue
            # continuation/variant row: data but no code
            if not code and items:
                v = {k: val for k, val in
                     [('model', model), ('cost', cost), ('labor', labor), ('description', desc)]
                     if val not in (None, '')}
                if v:
                    items[-1].setdefault('variants', []).append(v)
                continue
            if not code:
                continue

            item = {'code': code}
            if desc not in (None, ''):
                item['description'] = desc
            if model not in (None, ''):
                item['model'] = model
            if cost not in (None, ''):
                item['cost'] = cost
            if labor not in (None, ''):
                item['labor'] = labor
            if category:
                item['category'] = category
            if notes:
                item['notes'] = ' | '.join(str(n) for n in notes)
            item['section'] = SECTION.get(name, 'scope')
            item['sheet'] = name
            items.append(item)
        out[name] = items
        print(f"  {name:18} {len(items)} items")

    with open(OUT, 'w', encoding='utf-8') as f:
        yaml.safe_dump(out, f, allow_unicode=True, sort_keys=False, width=200)
    print(f"\nWrote {OUT}: {sum(len(v) for v in out.values())} codes total")


if __name__ == "__main__":
    main()
